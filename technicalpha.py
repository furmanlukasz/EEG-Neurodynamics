# import numpy as np
# import h5py
# from scipy.spatial.distance import euclidean, pdist, squareform
# import technic
# import pywt
# import matplotlib.pyplot as plt
# from pyrqa.image_generator import ImageGenerator
# from pyrqa.computation import RPComputation
# import numpy as np
# from pyrqa.time_series import TimeSeries
# from pyrqa.settings import Settings
# from pyrqa.analysis_type import Classic
# from pyrqa.neighbourhood import FixedRadius, RadiusCorridor
# from pyrqa.metric import EuclideanMetric
# from pyrqa.computation import RQAComputation
# from scipy import signal
# from nolitsa import delay, dimension
# import openpyxl


# def sublist(f):
#     '''tworzenie listy z osobami badanymi'''
#     struArray = f['ALLEEG']
#     subjects = []
#     for sb in range(len(struArray['subject'])):
#         subjects.append(''.join([chr(s) for s in list(f['#refs#'][struArray['subject'][sb][0]])]))

#     print(subjects)


# def builder(f, electro, subject):
#     """
#     returns specific data

#     first argument is a list of electrode labels

#     second argument is the name of the subject
#     """
#     struArray = f['ALLEEG']
#     ## lista z osobami
#     subjects = []
#     for sb in range(len(struArray['subject'])):
#         subjects.append(''.join([chr(s) for s in list(f['#refs#'][struArray['subject'][sb][0]])]))

#     ad = subjects.index(subject)
#     ## lista z adresami do elektrod
#     import re

#     pattern = '".*"'
#     '''
#     channelad = []
#     for ad in range(len(struArray['chanlocs'])):
#         match = re.search(pattern, str(f['#refs#'][struArray['chanlocs'][ad][0]])).group()[1:-1]
#         channelad.append(match)
#     ### adresy danych surowych
#     dataad = []
#     for a in range(len(struArray['uncuts'])):
#         macz = re.search(pattern, str(f['#refs#'][struArray['uncuts'][a][0]])).group()[1:-1]
#         dataad.append(macz)

#     sub_idx = subjects.index(subject)
#     sub_ad = channelad[sub_idx]
#     sb_data_ad = dataad[sub_idx]
#     ### dane ###
#     pierwszatabela = np.transpose(np.array(f[f[sb_data_ad]['data'][0][0]]))
#     '''

#     match = re.search(pattern, str(f['#refs#'][struArray['chanlocs'][ad][0]])).group()[1:-1]
#     macz = re.search(pattern, str(f['#refs#'][struArray['uncuts'][ad][0]])).group()[1:-1]

#     pierwszatabela = np.transpose(np.array(f[f[macz]['data'][0][0]]))
#     ### tworzenie listy elektrod
#     elektrody = []

#     for el in range(len(f[match]['labels'])):
#         elektrody.append(''.join([chr(s) for s in list(f[f[match]['labels'][el][0]][:])]))

#         ### wybór danych z elektrod podanych w zmiennej electro

#     dane = []
#     for ind in electro:
#         # print(ind)
#         # print(elektrody.index(ind))
#         wyn = []
#         x = list(pierwszatabela[elektrody.index(ind)][:])
#         dane.append(x)
#     return dane

# def rqacomp (f, title, stitle='test', neighber=0.6 , timedel=2, embedding=4, custom=False, save=True, xls=True, interval=0):
#     if not interval:
#         interval=len(f)
#     if not custom:
#         timedel = np.argmin(delay.dmi(f)).astype(np.int64)
#         dim = np.arange(1, 20)
#         print(dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2])
#         embedding = np.argmin(dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2]).astype(np.int64)
#         # embedding = np.argmin(fnns).astype(np.int64)
#     time_series = TimeSeries(f,
#                              embedding_dimension=embedding,
#                              time_delay=timedel)
#     settings = Settings(time_series,
#                             analysis_type=Classic,
#                             neighbourhood=FixedRadius(neighber),
#                             # neighbourhood=RadiusCorridor(inner_radius=0.32,outer_radius=0.86),
#                             similarity_measure=EuclideanMetric,
#                             theiler_corrector=1)
#         # computation = RQAComputation.create(settings,
#         #                                     verbose=True)
#         # result = computation.run()
#         # result.min_diagonal_line_length = 2
#         # result.min_vertical_line_length = 2
#         # result.min_white_vertical_line_length = 2
#     print(type(embedding), embedding)
#     print(type(timedel), timedel)
#     computation = RPComputation.create(settings)
#     result = computation.run()
#     dists = pdist(result.recurrence_matrix_reverse, metric='cosine')
#     plt.imshow(squareform(dists), cmap='jet', interpolation='nearest', origin='upper')
#     plt.colorbar()
#     ax = plt.gca()
#     ax.invert_yaxis()
#     plt.ylabel('ms')
#     plt.xlabel('ms')
#     # plt.xticks(range(min(0), max((interval/250)*100)))
#     # plt.yticks(range(min(0), max((interval/250)*100)))
#     plt.title(title +  'embedding = ' +str(embedding) + ', time delay = ' + str(timedel))
#     if save:
#         imag=stitle + '_embedding_' + str(embedding) + '.png'
#         plt.savefig(imag)
#         if xls:
#             wb = openpyxl.load_workbook('output.xlsx')
#             ws = wb.active
#             it=str(ws.max_row)
#             ws[str('A'+it)] = str(result.__str__())
#             img = openpyxl.drawing.image.Image(imag)
#             img.anchor(ws.cell('B'+it))
#             ws.add_image(img)
#             wb.save('output.xlsx')
#     else: plt.show()
#     plt.close()

# def fale(subject, dane, elnames, neighber=0.6 , timedel=2, embedding=4, custom=False, interval=None, save=True, xls=True):
#     if xls:
#         wb = openpyxl.Workbook()
#         wb.save('output.xlsx')
#     for el_idx in range(len(elnames)):
#         db4 = pywt.Wavelet('db4')
#         coeffs = pywt.wavedec(dane[el_idx], db4, mode='periodic', level=5)
#         dlu = len(dane[el_idx])
#         nazwy = ['delta', 'theta', 'alpha', 'beta', 'gamma', 'high gamma']
#         a = 0
#         # a=2
#         for i in coeffs:
#             f = signal.resample(i, dlu)
#             if interval:
#                 list_of_slices = zip(*(iter(f),) * interval)
#                 if not custom:
#                     dim = np.arange(1, 20)
#                     #fnns = dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2]
#                     embedding = np.argmin(dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2]).astype(np.int64)
#                 for j in list_of_slices:
#                     title=str('Electrode: ' + elnames[el_idx] + ', ' + nazwy[a] + ' band, ' + ' timestamp ' + str(interval)  )
#                     stitle=str(subject + '_' + elnames[el_idx] + '_' + nazwy[a] +' timestamp ' + str(interval) )
#                     rqacomp(f, title, stitle, neighber, timedel, embedding, custom, save, xls)
#             else:
#                 title = str(
#                     'Electrode: ' + elnames[el_idx] + ', ' + nazwy[a] + ' band, ' )
#                 stitle = str(subject + '_' + elnames[el_idx] + '_' + nazwy[a] )
#                 if not custom:
#                     timedel = np.argmin(delay.dmi(f)).astype(np.int64)
#                     dim = np.arange(1, 20)
#                     fnns = dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2]
#                     embedding = np.argmin(fnns).astype(np.int64)
#                 rqacomp(f, title, stitle, neighber, timedel, embedding, custom, save, xls)
#         a = a + 1





# import numpy as np
# import h5py
# from scipy.spatial.distance import euclidean, pdist, squareform
# import technic
# import pywt
# import matplotlib.pyplot as plt
# from pyrqa.image_generator import ImageGenerator
# from pyrqa.computation import RPComputation
# import numpy as np
# from pyrqa.time_series import TimeSeries
# from pyrqa.settings import Settings
# from pyrqa.analysis_type import Classic
# from pyrqa.neighbourhood import FixedRadius, RadiusCorridor
# from pyrqa.metric import EuclideanMetric
# from pyrqa.computation import RQAComputation
# from pyrqa.neighbourhood import Unthresholded
# from scipy import signal
# from nolitsa import delay, dimension
# import openpyxl


# def sublist(f):
#     '''tworzenie listy z osobami badanymi'''
#     struArray = f['ALLEEG']
#     subjects = []
#     for sb in range(len(struArray['subject'])):
#         subjects.append(''.join([chr(s) for s in list(f['#refs#'][struArray['subject'][sb][0]])]))

#     print(subjects)


# def builder(f, electro, subject):
#     """
#     returns specific data

#     first argument is a list of electrode labels

#     second argument is the name of the subject
#     """
#     struArray = f['ALLEEG']
#     ## lista z osobami
#     subjects = []
#     for sb in range(len(struArray['subject'])):
#         subjects.append(''.join([chr(s) for s in list(f['#refs#'][struArray['subject'][sb][0]])]))

#     ad = subjects.index(subject)
#     ## lista z adresami do elektrod
#     import re

#     pattern = '".*"'
#     '''
#     channelad = []
#     for ad in range(len(struArray['chanlocs'])):
#         match = re.search(pattern, str(f['#refs#'][struArray['chanlocs'][ad][0]])).group()[1:-1]
#         channelad.append(match)
#     ### adresy danych surowych
#     dataad = []
#     for a in range(len(struArray['uncuts'])):
#         macz = re.search(pattern, str(f['#refs#'][struArray['uncuts'][a][0]])).group()[1:-1]
#         dataad.append(macz)

#     sub_idx = subjects.index(subject)
#     sub_ad = channelad[sub_idx]
#     sb_data_ad = dataad[sub_idx]
#     ### dane ###
#     pierwszatabela = np.transpose(np.array(f[f[sb_data_ad]['data'][0][0]]))
#     '''

#     match = re.search(pattern, str(f['#refs#'][struArray['chanlocs'][ad][0]])).group()[1:-1]
#     macz = re.search(pattern, str(f['#refs#'][struArray['uncuts'][ad][0]])).group()[1:-1]

#     pierwszatabela = np.transpose(np.array(f[f[macz]['data'][0][0]]))
#     ### tworzenie listy elektrod
#     elektrody = []

#     for el in range(len(f[match]['labels'])):
#         elektrody.append(''.join([chr(s) for s in list(f[f[match]['labels'][el][0]][:])]))

#         ### wybór danych z elektrod podanych w zmiennej electro

#     dane = []
#     for ind in electro:
#         # print(ind)
#         # print(elektrody.index(ind))
#         wyn = []
#         x = list(pierwszatabela[elektrody.index(ind)][:])
#         dane.append(x)
#     return dane

# def rqacomp (f, title, stitle='test', neighber=0.6 , timedel=2, embedding=4, custom=False, save=True, xls=True, interval=0):
#     if not interval:
#         interval=len(f)
#     if not custom:
#         timedel = np.argmin(delay.dmi(f))
#         dim = np.arange(1, 20)
#         # fnns = dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2]
#         # embedding = np.argmin(fnns)
#         embedding = np.argmin(dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2])
#     time_series = TimeSeries(f,
#                              embedding_dimension=embedding,
#                              time_delay=timedel)
#     settings = Settings(time_series,
#                         analysis_type=Classic,
#                         neighbourhood=FixedRadius(neighber),
#                         #neighbourhood=Unthresholded(),
#                         #neighbourhood=RadiusCorridor(inner_radius=0.32,outer_radius=0.86),
#                         similarity_measure=EuclideanMetric,
#                         theiler_corrector=1)
#     computation = RQAComputation.create(settings,
#                                          verbose=True)
#     result = computation.run()
#     result.min_diagonal_line_length = 2
#     result.min_vertical_line_length = 2
#     result.min_white_vertical_line_length = 2
#     entryArray=result.to_array()
#     settings = Settings(time_series,
#                         analysis_type=Classic,
#                         #neighbourhood=FixedRadius(neighber),
#                         neighbourhood=Unthresholded(),
#                         # neighbourhood=RadiusCorridor(inner_radius=0.32,outer_radius=0.86),
#                         similarity_measure=EuclideanMetric,
#                         theiler_corrector=1)
#     computation = RPComputation.create(settings)
#     result = computation.run()
#     dists = pdist(result.recurrence_matrix_reverse, metric='cosine')
#     plt.imshow(squareform(dists), cmap='jet', interpolation='nearest', origin='upper')
#     plt.colorbar()
#     ax = plt.gca()
#     ax.invert_yaxis()
#     plt.ylabel('ms')
#     plt.xlabel('ms')
#     plt.xlim(0, int((interval/250)*100))
#     plt.ylim(0, int((interval/250)*100))
#     plt.title(title +  'embedding = ' +str(embedding) + ', time delay = ' + str(timedel))
#     if save:
#         imag=stitle + '_embedding_' + str(embedding) + '.png'
#         plt.savefig(imag)
#         if xls:
#             wb = openpyxl.load_workbook('output.xlsx')
#             ws = wb.active
#             it=str(ws.max_row)
#             ws[str('A'+it)] = str(result.__str__())

#             #for b in range(len(entryArray)):
#             #    temp = entryArray[b]
#             #    ws.cell(row=ws.max_row+b, column=1).value = temp
#             img = openpyxl.drawing.image.Image(imag)
#             img.anchor = str('B'+str(it))
#             ws.add_image(img)
#             wb.save('output.xlsx')
#     else: plt.show()
#     plt.close()

# def fale(subject, dane, elnames, neighber=0.6 , timedel=2, embedding=4, custom=False, interval=None, save=True, xls=True):
#     if xls:
#         wb = openpyxl.Workbook()
#         wb.save('output.xlsx')
#     for el_idx in range(len(elnames)):
#         db4 = pywt.Wavelet('db4')
#         coeffs = pywt.wavedec(dane[el_idx], db4, mode='periodic', level=5)
#         dlu = len(dane[el_idx])
#         nazwy = ['delta', 'theta', 'alpha', 'beta', 'gamma', 'high gamma']
#         a = 0
#         # a=2
#         for i in coeffs:
#             f = signal.resample(i, dlu)
#             if interval:
#                 list_of_slices = zip(*(iter(f),) * interval)
#                 if not custom:
#                     dim = np.arange(1, 20)
#                     # fnns = dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2]
#                     # embedding = np.argmin(fnns)
#                     embedding = np.argmin(dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2])
#                 for j in list_of_slices:
#                     title=str('Electrode: ' + elnames[el_idx] + ', ' + nazwy[a] + ' band, ' + ' timestamp ' + str(interval)  )
#                     stitle=str(subject + '_' + elnames[el_idx] + '_' + nazwy[a] +' timestamp ' + str(interval) )
#                     rqacomp(f, title, stitle, neighber, timedel, embedding, custom, save, xls)
#             else:
#                 title = str(
#                     'Electrode: ' + elnames[el_idx] + ', ' + nazwy[a] + ' band, ' )
#                 stitle = str(subject + '_' + elnames[el_idx] + '_' + nazwy[a] )
#                 if not custom:
#                     timedel = np.argmin(delay.dmi(f))
#                     dim = np.arange(1, 20)
#                     # fnns = dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2]
#                     # embedding = np.argmin(fnns)
#                     embedding = np.argmin(dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2])
#                 rqacomp(f, title, stitle, neighber, timedel, embedding, custom, save, xls)
#         a = a + 1





import numpy as np
import h5py
from scipy.spatial.distance import euclidean, pdist, squareform
import technic
import pywt
import matplotlib.pyplot as plt
from pyrqa.image_generator import ImageGenerator
from pyrqa.computation import RPComputation
import numpy as np
from pyrqa.time_series import TimeSeries
from pyrqa.settings import Settings
from pyrqa.analysis_type import Classic
from pyrqa.neighbourhood import FixedRadius, RadiusCorridor
from pyrqa.metric import EuclideanMetric
from pyrqa.computation import RQAComputation
from pyrqa.neighbourhood import Unthresholded
from scipy import signal
from nolitsa import delay, dimension
import openpyxl


def sublist(f):
    '''tworzenie listy z osobami badanymi'''
    struArray = f['ALLEEG']
    subjects = []
    for sb in range(len(struArray['subject'])):
        subjects.append(''.join([chr(s) for s in list(f['#refs#'][struArray['subject'][sb][0]])]))

    print(subjects)


def builder(f, electro, subject):
    """
    returns specific data

    first argument is a list of electrode labels

    second argument is the name of the subject
    """
    struArray = f['ALLEEG']
    ## lista z osobami
    subjects = []
    for sb in range(len(struArray['subject'])):
        subjects.append(''.join([chr(s) for s in list(f['#refs#'][struArray['subject'][sb][0]])]))

    ad = subjects.index(subject)
    ## lista z adresami do elektrod
    import re

    pattern = '".*"'
    '''
    channelad = []
    for ad in range(len(struArray['chanlocs'])):
        match = re.search(pattern, str(f['#refs#'][struArray['chanlocs'][ad][0]])).group()[1:-1]
        channelad.append(match)
    ### adresy danych surowych
    dataad = []
    for a in range(len(struArray['uncuts'])):
        macz = re.search(pattern, str(f['#refs#'][struArray['uncuts'][a][0]])).group()[1:-1]
        dataad.append(macz)

    sub_idx = subjects.index(subject)
    sub_ad = channelad[sub_idx]
    sb_data_ad = dataad[sub_idx]
    ### dane ###
    pierwszatabela = np.transpose(np.array(f[f[sb_data_ad]['data'][0][0]]))
    '''

    match = re.search(pattern, str(f['#refs#'][struArray['chanlocs'][ad][0]])).group()[1:-1]
    macz = re.search(pattern, str(f['#refs#'][struArray['uncuts'][ad][0]])).group()[1:-1]

    pierwszatabela = np.transpose(np.array(f[f[macz]['data'][0][0]]))
    ### tworzenie listy elektrod
    elektrody = []

    for el in range(len(f[match]['labels'])):
        elektrody.append(''.join([chr(s) for s in list(f[f[match]['labels'][el][0]][:])]))

        ### wybór danych z elektrod podanych w zmiennej electro

    dane = []
    for ind in electro:
        # print(ind)
        # print(elektrody.index(ind))
        wyn = []
        x = list(pierwszatabela[elektrody.index(ind)][:])
        dane.append(x)
    return dane


def rqacomp (f, title, stitle='test', neighber=0.6 , timedel=2, embedding=4, custom=False, save=True, xls=True, interval=0):
    if not interval:
        interval=len(f)
    # if not custom:
    #     timedel = np.argmin(delay.dmi(f)[:26])
    #     dim = np.arange(1, 20)
    #     # fnns = dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2]
    #     # embedding = np.argmin(fnns)
    #     embedding = np.argmin(dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2])
    time_series = TimeSeries(f,
                             embedding_dimension=embedding,
                             time_delay=timedel)
    settings = Settings(time_series,
                        analysis_type=Classic,
                        neighbourhood=FixedRadius(neighber),
                        #neighbourhood=Unthresholded(),
                        #neighbourhood=RadiusCorridor(inner_radius=0.32,outer_radius=0.86),
                        similarity_measure=EuclideanMetric,
                        theiler_corrector=1)
    computation = RQAComputation.create(settings,
                                         verbose=True)
    result = computation.run()
    result.min_diagonal_line_length = 2
    result.min_vertical_line_length = 2
    result.min_white_vertical_line_length = 2
    entryArray=result.to_array()
    tt=entryArray[10]
    settings = Settings(time_series,
                        analysis_type=Classic,
                        #neighbourhood=FixedRadius(neighber),
                        neighbourhood=Unthresholded(),
                        # neighbourhood=RadiusCorridor(inner_radius=0.32,outer_radius=0.86),
                        similarity_measure=EuclideanMetric,
                        theiler_corrector=1)
    computation = RPComputation.create(settings)
    result = computation.run()
    dists = pdist(result.recurrence_matrix_reverse, metric='cosine')
    '''
    plt.imshow(squareform(dists), cmap='jet', interpolation='nearest', origin='upper')
    plt.colorbar()
    ax = plt.gca()
    ax.invert_yaxis()
    plt.ylabel('ms')
    plt.xlabel('ms')
    plt.xlim(0, int((interval/250)*100))
    plt.ylim(0, int((interval/250)*100))
    plt.title(title +  'embedding = ' +str(embedding) + ', time delay = ' + str(timedel))
    '''
    if save:
        imag=stitle + '_embedding_' + str(embedding) + '.png'
        plt.savefig(imag)
        if xls:
            wb = openpyxl.load_workbook('output.xlsx')
            ws = wb.active
            it=str(ws.max_row+30)
            print(it)
            ws[str('A'+it)] = str(result.__str__())
            # do zrobienia
            for b in range(len(entryArray)):
                temp = entryArray[b]
                vname=['Min_DLL', 'min_VLL', 'min_WVLL', 'RR', 'DET', 'ADL', 'LDL', 'DIV', 'EDL', 'LAM', 'TT', 'LVL', 'EVL', 'AWVL', 'LWVL', 'LWVLI', 'EWVL', 'Rat_Det/RR', 'Rat_Lam/Det']
                row=ws.max_row+1
                ws.cell(row=row, column=12).value = vname[b]
                ws.cell(row=row, column=13).value = temp
            img = openpyxl.drawing.image.Image(imag)
            img.anchor = str('C'+str(it))
            ws.add_image(img)
            wb.save('output.xlsx')
    # else: plt.show()
    # plt.close()
    return tt
    

def fale(subject, dane, elnames, neighber=0.6 , timedel=2, embedding=4, custom=False, interval=None, save=True, xls=True, custom_embedding=False):
    
    if xls:
        wb = openpyxl.Workbook()
        wb.save('output.xlsx')
    for el_idx in range(len(elnames)):
        db4 = pywt.Wavelet('db4')
        coeffs = pywt.wavedec(dane[el_idx], db4, mode='periodic', level=5)
        dlu = len(dane[el_idx])
        nazwy = ['delta', 'theta', 'alpha', 'beta', 'gamma', 'high gamma']
        a = 0
        # a=2
        # all_embs=[]
        # all_tt=[]
        tt_dic={}
        for i in coeffs:
            band_embs=[]
            band_tt=[]
            band_fnns=[]
            f = signal.resample(i, dlu)
            
            if interval:
                list_of_slices = zip(*(iter(f),) * interval)
                # if not custom:
                #     timedel = np.argmin(delay.dmi(f)[:26])
                #     dim = np.arange(1, 20)
                #     # fnns = dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2]
                #     # embedding = np.argmin(fnns)
                #     embedding = np.argmin(dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2])
                   
                counter=1
                for j in list_of_slices:
                    if not custom:
                        timedel = np.argmin(delay.dmi(j)[:26])
                        dim = np.arange(1, 20)
                        # fnns = dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2]
                        # embedding = np.argmin(fnns)
                        fnn_min=np.amin(dimension.fnn(j, dim=dim, tau=timedel, metric='euclidean')[2])
                        embedding = np.argmin(dimension.fnn(j, dim=dim, tau=timedel, metric='euclidean')[2])
                    title=str('Electrode: ' + elnames[el_idx] + ', ' + nazwy[a] + ' band, ' + ' timestamp ' + str((interval*counter)/250)  )
                    stitle=str(subject + '_' + elnames[el_idx] + '_' + nazwy[a] +' timestamp ' + str(interval) )
                    tt=rqacomp(j, title, stitle, neighber, timedel, embedding, custom, save, xls, interval)
                    
                    band_embs.append(embedding)
                    band_fnns.append(fnn_min)
                    band_tt.append(tt)
                    band_tt=np.nan_to_num(band_tt, copy=False, nan=0).tolist()
                    print(band_embs)
                    print(band_tt)
                    counter+=1
                plt.scatter(band_tt, band_fnns)
                plt.title('Electrode: ' + elnames[el_idx] + ' ' + nazwy[a] + ' band')
                plt.ylabel("Minimum FNNs")
                plt.xlabel('Trapping times')
                plt.savefig( elnames[el_idx] +'_' + nazwy[a] + '_fnns_tt.png') 
                plt.close()
                    
               
                # np.nan_to_num(all_tt, copy=False, nan=0)
               
            else:
                title = str(
                    'Electrode: ' + elnames[el_idx] + ', ' + nazwy[a] + ' band, ' )
                stitle = str(subject + '_' + elnames[el_idx] + '_' + nazwy[a] )
                
                # if custom_embedding:
                #     for emb in embedding:
                #         tt=rqacomp(f, title, stitle, neighber, timedel, emb, custom, save, xls, interval)

                #         band_tt.append(tt)
                #         band_tt=np.nan_to_num(band_tt, copy=False, nan=0).tolist()

                
                
                
                if not custom:
                    timedel = np.argmin(delay.dmi(f)[:26])
                    dim = np.arange(1, 20)
                    # fnns = dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2]
                    # embedding = np.argmin(fnns)
                    print(timedel)
                     
                    embedding = np.argmin(dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2])
                
                # fnn_min=np.amin(dimension.fnn(f, dim=dim, tau=timedel, metric='euclidean')[2])
                tt=rqacomp(f, title, stitle, neighber, timedel, embedding, custom, save, xls, interval)
                

                # all_embs.append(embedding)
                # all_tt.append(tt)
              
    
            # plt.plot(all_embs, all_tt)
            # plt.savefig(nazwy[a] + 'embs_tt.png')  
            # tt_dic[nazwy[a]]=band_tt  
            a = a + 1
            # all_tt=np.nan_to_num(all_tt, copy=False, nan=0)
            # plt.scatter(all_tt, all_embs)
            # plt.title('Electrode: ' + elnames[el_idx] + ', ' + ' all bands ')
            # plt.ylabel("Embeddings")
            # plt.xlabel('Trapping times')
            # plt.savefig( elnames[el_idx] + '_' + 'embs_tt.png') 
            # plt.close()
    return tt_dic            
    # print(all_embs)
    # print(all_tt)
    # plt.scatter(all_embs, all_tt)
    # plt.xlabel("Embeddings")
    # plt.ylabel('Trapping times')
    # plt.savefig(nazwy[a] + 'embs_tt.png') 
     

def fnny(subject, dane, elnames, embedding, metric='euclidean'):
    all_fnn_dic={}
    for el_idx in range(len(elnames)):
        db4 = pywt.Wavelet('db4')
        coeffs = pywt.wavedec(dane[el_idx], db4, mode='periodic', level=5)
        dlu = len(dane[el_idx])
        nazwy = ['delta', 'theta', 'alpha', 'beta', 'gamma', 'high gamma']
        a = 0

        fnn_dic={}
        for i in coeffs:
            f = signal.resample(i, dlu)
            timedel = np.argmin(delay.dmi(f)[:26])
            print(timedel)
            fnns = dimension.fnn(f, dim=embedding, tau=timedel, metric=metric)[2]

            fnn_dic[nazwy[a]]=fnns
            a=a+1
        all_fnn_dic[elnames[el_idx]]=fnn_dic
    return all_fnn_dic



def tts(subject, dane, elnames, embedding, neighber=0.6):
    all_tt_dic={}
    for el_idx in range(len(elnames)):
        db4 = pywt.Wavelet('db4')
        coeffs = pywt.wavedec(dane[el_idx], db4, mode='periodic', level=5)
        dlu = len(dane[el_idx])
        nazwy = ['delta', 'theta', 'alpha', 'beta', 'gamma', 'high gamma']
        a = 0

        tt_dic={}
        
        for i in coeffs:
            band_tt=[]
            f = signal.resample(i, dlu)
            title = str(
            'Electrode: ' + elnames[el_idx] + ', ' + nazwy[a] + ' band, ' )
            stitle = str(subject + '_' + elnames[el_idx] + '_' + nazwy[a] )
            for emb in embedding:
                timedel = np.argmin(delay.dmi(f)[:26])
                tt=rqacomp(f, title, stitle, neighber=neighber, timedel=timedel, embedding=emb, custom=True, save=False, xls=False, interval=None)

                band_tt.append(tt)
                # band_tt=np.nan_to_num(band_tt, copy=False, nan=0).tolist()
                print(band_tt)
            
            tt_dic[nazwy[a]]=band_tt
            a=a+1
        all_tt_dic[elnames[el_idx]]=tt_dic
    return all_tt_dic

def timedels(subject, dane, elnames):
    all_timedel_dic={}
    for el_idx in range(len(elnames)):
        db4 = pywt.Wavelet('db4')
        coeffs = pywt.wavedec(dane[el_idx], db4, mode='periodic', level=5)
        dlu = len(dane[el_idx])
        nazwy = ['delta', 'theta', 'alpha', 'beta', 'gamma', 'high gamma']
        a = 0

        timedel_dic={}
        for i in coeffs:
            f = signal.resample(i, dlu)
            timedel = np.argmin(delay.dmi(f)[:26])
            

            timedel_dic[nazwy[a]]=timedel
            a=a+1
        all_timedel_dic[elnames[el_idx]]=timedel_dic
    return all_timedel_dic
   






####
